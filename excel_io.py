from paths import input_path, output_path
from datetime import date, datetime
from openpyxl import Workbook, load_workbook


def format_date(value):
    """Convert a worksheet date or string into ISO format for the API."""
    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, str):
        value = value.strip()
        if not value:
            raise ValueError("empty date string")
        if value.count("-") == 2:
            return value
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue

    raise ValueError(f"Unsupported start date format: {value!r}")


def load_dockets():
    """Load docket numbers and start dates from the Excel worksheet."""
    workbook = load_workbook(input_path)
    worksheet = workbook.active
    docket_rows = []

    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=2):
        docket_cell, start_date_cell = row
        docket = docket_cell.value
        if docket is None:
            continue

        docket = str(docket).strip()
        if not docket:
            continue

        docket_rows.append({
            "docket": docket,
            "start_date": start_date_cell.value,
            "start_date_cell": start_date_cell,
        })

    return workbook, docket_rows


def load_results():
    """Load the results workbook, clear non-TEMPLATE sheets, and verify TEMPLATE exists."""

    workbook = load_workbook(output_path)

    if "TEMPLATE" not in workbook.sheetnames:
        raise ValueError(f"{output_path} must contain a sheet named 'TEMPLATE'.")

    for sheet_name in list(workbook.sheetnames):
        if sheet_name != "TEMPLATE":
            workbook.remove(workbook[sheet_name])

    return workbook


def create_result_sheet(workbook, docket, records, template_name="TEMPLATE"):
    """Create or replace a docket sheet left of TEMPLATE and populate it."""
    if docket in workbook.sheetnames:
        workbook.remove(workbook[docket])

    template_index = workbook.sheetnames.index(template_name)
    worksheet = workbook.create_sheet(title=docket, index=template_index)

    template = workbook[template_name]
    headers = [cell for cell in next(template.iter_rows(min_row=1, max_row=1, values_only=True))]
    if not headers:
        headers = ["title", "date", "category", "accession_number", "accession_url"]

    worksheet.append(headers)
    for record in records:
        worksheet.append([
            record.get("title", ""),
            record.get("date", ""),
            record.get("category", ""),
            record.get("accession_number", ""),
            record.get("accession_url", ""),
        ])
